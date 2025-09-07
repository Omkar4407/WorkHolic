from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'

# Absolute path for PythonAnywhere
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_FILE = os.path.join(DATA_DIR, "employees.xlsx")

# --------------------- Time Utilities ---------------------
def get_current_time():
    """Return local time (e.g., IST)"""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)  # Adjust for your timezone

def format_time(dt):
    """Return formatted time string"""
    return dt.strftime('%I:%M %p')

# --------------------- Initialize Excel ---------------------
def init_excel_file():
    """Initialize Excel file with Users and Attendance sheets"""
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)
    
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        
        # Users sheet
        users_sheet = wb.active
        users_sheet.title = 'Users'
        users_sheet.append(['Username', 'Password', 'Role'])
        users_sheet.append(['omkar', 'admin123', 'Admin'])
        users_sheet.append(['john', 'john123', 'Employee'])
        users_sheet.append(['sarah', 'sarah123', 'Employee'])
        
        # Attendance sheet
        attendance_sheet = wb.create_sheet('Attendance')
        attendance_sheet.append(['Date', 'Username', 'Clock-In', 'Clock-Out', 'Break Start', 'Break End', 'Total Hours'])
        attendance_sheet.append(['2025-01-07', 'john', '09:05 AM', '06:15 PM', '01:00 PM', '01:30 PM', '8:40'])
        
        wb.save(EXCEL_FILE)

# --------------------- User & Attendance Functions ---------------------
def load_users():
    wb = load_workbook(EXCEL_FILE)
    users_sheet = wb['Users']
    users = {}
    for row in users_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            users[row[0]] = {'password': row[1], 'role': row[2]}
    return users

def save_user(username, password, role):
    wb = load_workbook(EXCEL_FILE)
    users_sheet = wb['Users']
    user_exists = False
    for row in users_sheet.iter_rows(min_row=2):
        if row[0].value == username:
            row[1].value = password
            row[2].value = role
            user_exists = True
            break
    if not user_exists:
        users_sheet.append([username, password, role])
    wb.save(EXCEL_FILE)

def delete_user(username):
    wb = load_workbook(EXCEL_FILE)
    users_sheet = wb['Users']
    for row in users_sheet.iter_rows(min_row=2):
        if row[0].value == username:
            users_sheet.delete_rows(row[0].row)
            break
    wb.save(EXCEL_FILE)

def load_attendance(username=None, start_date=None, end_date=None):
    wb = load_workbook(EXCEL_FILE)
    attendance_sheet = wb['Attendance']
    records = []
    for row in attendance_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            record = {
                'date': row[0],
                'username': row[1],
                'clock_in': row[2],
                'clock_out': row[3],
                'break_start': row[4],
                'break_end': row[5],
                'total_hours': row[6]
            }
            if username and record['username'] != username:
                continue
            if start_date and str(record['date']) < start_date:
                continue
            if end_date and str(record['date']) > end_date:
                continue
            records.append(record)
    return records

def save_attendance_record(username, action, timestamp):
    wb = load_workbook(EXCEL_FILE)
    attendance_sheet = wb['Attendance']
    today = timestamp.strftime('%Y-%m-%d')
    time_str = format_time(timestamp)
    
    record_row = None
    for row in attendance_sheet.iter_rows(min_row=2):
        if str(row[0].value) == today and row[1].value == username:
            record_row = row
            break
    
    if not record_row:
        attendance_sheet.append([today, username, '', '', '', '', ''])
        record_row = list(attendance_sheet.iter_rows())[-1]
    
    # Record action
    if action == 'clock_in':
        record_row[2].value = time_str
    elif action == 'clock_out':
        record_row[3].value = time_str
    elif action == 'break_start':
        record_row[4].value = time_str
    elif action == 'break_end':
        record_row[5].value = time_str
    
    # Calculate total hours if clock-in and clock-out exist
    if record_row[2].value and record_row[3].value:
        clock_in = datetime.strptime(record_row[2].value, '%I:%M %p')
        clock_out = datetime.strptime(record_row[3].value, '%I:%M %p')
        work_duration = clock_out - clock_in
        
        # Subtract break if exists
        if record_row[4].value and record_row[5].value:
            break_start = datetime.strptime(record_row[4].value, '%I:%M %p')
            break_end = datetime.strptime(record_row[5].value, '%I:%M %p')
            work_duration -= (break_end - break_start)
        
        total_seconds = work_duration.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        record_row[6].value = f"{hours}:{minutes:02d}"
    
    wb.save(EXCEL_FILE)

def get_today_status(username):
    today = get_current_time().strftime('%Y-%m-%d')
    records = load_attendance(username=username)
    today_record = None
    for record in records:
        if str(record['date']) == today:
            today_record = record
            break
    if not today_record:
        return {'clocked_in': False, 'on_break': False, 'clocked_out': False}
    return {
        'clocked_in': bool(today_record['clock_in']),
        'on_break': bool(today_record['break_start'] and not today_record['break_end']),
        'clocked_out': bool(today_record['clock_out'])
    }

# --------------------- Routes ---------------------
@app.route('/')
def index():
    if 'username' in session:
        if session['role'] == 'Admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = load_users()
        if username in users and users[username]['password'] == password:
            session['username'] = username
            session['role'] = users[username]['role']
            if users[username]['role'] == 'Admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('employee_dashboard'))
        else:
            flash('Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/employee')
def employee_dashboard():
    if 'username' not in session or session['role'] != 'Employee':
        return redirect(url_for('login'))
    username = session['username']
    status = get_today_status(username)
    records = load_attendance(username=username)
    records.sort(key=lambda x: str(x['date']), reverse=True)
    return render_template('employee_dashboard.html', username=username, status=status, records=records[:10])

@app.route('/admin')
def admin_dashboard():
    if 'username' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    users = load_users()
    records = load_attendance()
    records.sort(key=lambda x: str(x['date']), reverse=True)
    return render_template('admin_dashboard.html', users=users, records=records[:20])

@app.route('/time_action', methods=['POST'])
def time_action():
    if 'username' not in session or session['role'] != 'Employee':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    action = request.json.get('action')
    username = session['username']
    if action not in ['clock_in', 'clock_out', 'break_start', 'break_end']:
        return jsonify({'success': False, 'message': 'Invalid action'})
    try:
        timestamp = get_current_time()
        save_attendance_record(username, action, timestamp)
        return jsonify({'success': True, 'message': f'{action.replace("_", " ").title()} recorded successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/manage_user', methods=['POST'])
def manage_user():
    if 'username' not in session or session['role'] != 'Admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    action = request.json.get('action')
    username = request.json.get('username')
    password = request.json.get('password')
    role = request.json.get('role')
    try:
        if action == 'add' or action == 'edit':
            save_user(username, password, role)
            return jsonify({'success': True, 'message': f'User {action}ed successfully'})
        elif action == 'delete':
            delete_user(username)
            return jsonify({'success': True, 'message': 'User deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Invalid action'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/filter_attendance', methods=['POST'])
def filter_attendance():
    if 'username' not in session or session['role'] != 'Admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    username = request.json.get('username')
    start_date = request.json.get('start_date')
    end_date = request.json.get('end_date')
    records = load_attendance(username=username, start_date=start_date, end_date=end_date)
    records.sort(key=lambda x: str(x['date']), reverse=True)
    return jsonify({'success': True, 'records': records})

@app.route('/export_attendance')
def export_attendance():
    if 'username' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    records = load_attendance()
    output = BytesIO()
    csv_data = [['Date', 'Username', 'Clock-In', 'Clock-Out', 'Break Start', 'Break End', 'Total Hours']]
    for record in records:
        csv_data.append([
            str(record['date']),
            record['username'],
            record['clock_in'] or '',
            record['clock_out'] or '',
            record['break_start'] or '',
            record['break_end'] or '',
            record['total_hours'] or ''
        ])
    csv_string = '\n'.join([','.join(row) for row in csv_data])
    output.write(csv_string.encode())
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue()),
        mimetype='text/csv',
        as_attachment=True,
        download_name='attendance_export.csv'
    )

# --------------------- Init ---------------------
init_excel_file()
# app.run() is handled by PythonAnywhere WSGI
