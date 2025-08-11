import os
from openpyxl import load_workbook, Workbook
from PIL import Image
import numpy as np
import face_recognition
from data_logger import ensure_excel_file

def register_user(data, image):
    ensure_excel_file()
    emp_id = data.get('employeeId')
    name = data.get('fullName')
    dept = data.get('department')
    img_dir = os.path.join(os.path.dirname(__file__), '../public/images')
    img_path = os.path.join(img_dir, f'{emp_id}.jpg')

    # Save image
    image.save(img_path)

    # Check for duplicate EmployeeID
    wb = load_workbook('data/work_log.xlsx')
    ws = wb['Registered_Users']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == emp_id:
            return {'success': False, 'message': 'Employee ID already registered.'}

    # Check if face is present
    img = face_recognition.load_image_file(img_path)
    faces = face_recognition.face_encodings(img)
    if not faces:
        os.remove(img_path)
        return {'success': False, 'message': 'No face detected in image.'}

    # Append to Excel
    ws.append([emp_id, name, dept, f'public/images/{emp_id}.jpg'])
    wb.save('data/work_log.xlsx')
    return {'success': True, 'message': 'Registration successful.'}
