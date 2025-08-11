import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def ensure_excel_file():
    path = 'data/work_log.xlsx'
    if not os.path.exists(path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Registered_Users'
        ws1.append(['EmployeeID', 'Name', 'Department', 'Image Path'])
        ws2 = wb.create_sheet('Work_Log')
        ws2.append(['EmployeeID', 'Name', 'Date', 'Login Time', 'Breaks', 'Resumes', 'Active Time', 'Break Time', 'Logoff Time'])
        wb.save(path)
    else:
        wb = load_workbook(path)
        if 'Registered_Users' not in wb.sheetnames:
            ws1 = wb.create_sheet('Registered_Users')
            ws1.append(['EmployeeID', 'Name', 'Department', 'Image Path'])
        if 'Work_Log' not in wb.sheetnames:
            ws2 = wb.create_sheet('Work_Log')
            ws2.append(['EmployeeID', 'Name', 'Date', 'Login Time', 'Breaks', 'Resumes', 'Active Time', 'Break Time', 'Logoff Time'])
        wb.save(path)

def get_registered_users():
    wb = load_workbook('data/work_log.xlsx')
    ws = wb['Registered_Users']
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        users.append({'employeeId': row[0], 'name': row[1], 'department': row[2], 'imagePath': row[3]})
    return users

def log_login(emp_id):
    wb = load_workbook('data/work_log.xlsx')
    ws_users = wb['Registered_Users']
    ws_log = wb['Work_Log']
    name = None
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == emp_id:
            name = row[1]
            break
    if not name:
        return
    today = datetime.now().strftime('%Y-%m-%d')
    for row in ws_log.iter_rows(min_row=2):
        if row[0].value == emp_id and row[2].value == today:
            if not row[3].value:
                row[3].value = datetime.now().strftime('%H:%M:%S')
                wb.save('data/work_log.xlsx')
            return
    ws_log.append([emp_id, name, today, datetime.now().strftime('%H:%M:%S'), '', '', '', '', ''])
    wb.save('data/work_log.xlsx')

def log_break(emp_id):
    wb = load_workbook('data/work_log.xlsx')
    ws_log = wb['Work_Log']
    today = datetime.now().strftime('%Y-%m-%d')
    for row in ws_log.iter_rows(min_row=2):
        if row[0].value == emp_id and row[2].value == today:
            breaks = row[4].value or ''
            now = datetime.now().strftime('%H:%M:%S')
            row[4].value = (breaks + ',' if breaks else '') + now
            wb.save('data/work_log.xlsx')
            return

def log_resume(emp_id):
    wb = load_workbook('data/work_log.xlsx')
    ws_log = wb['Work_Log']
    today = datetime.now().strftime('%Y-%m-%d')
    for row in ws_log.iter_rows(min_row=2):
        if row[0].value == emp_id and row[2].value == today:
            resumes = row[5].value or ''
            now = datetime.now().strftime('%H:%M:%S')
            row[5].value = (resumes + ',' if resumes else '') + now
            wb.save('data/work_log.xlsx')
            return

def log_logoff(emp_id):
    wb = load_workbook('data/work_log.xlsx')
    ws_log = wb['Work_Log']
    today = datetime.now().strftime('%Y-%m-%d')
    for row in ws_log.iter_rows(min_row=2):
        if row[0].value == emp_id and row[2].value == today:
            if not row[8].value:
                row[8].value = datetime.now().strftime('%H:%M:%S')
                # Calculate active and break time
                login = row[3].value
                breaks = (row[4].value or '').split(',') if row[4].value else []
                resumes = (row[5].value or '').split(',') if row[5].value else []
                logoff = row[8].value
                # Calculate times
                fmt = '%H:%M:%S'
                total_active = 0
                total_break = 0
                if login and logoff:
                    login_dt = datetime.strptime(login, fmt)
                    logoff_dt = datetime.strptime(logoff, fmt)
                    total = (logoff_dt - login_dt).total_seconds()
                    # Pair breaks and resumes
                    for b, r in zip(breaks, resumes):
                        b_dt = datetime.strptime(b, fmt)
                        r_dt = datetime.strptime(r, fmt)
                        total_break += (r_dt - b_dt).total_seconds()
                    # If more breaks than resumes, assume last break until logoff
                    if len(breaks) > len(resumes):
                        b_dt = datetime.strptime(breaks[-1], fmt)
                        r_dt = logoff_dt
                        total_break += (r_dt - b_dt).total_seconds()
                    total_active = total - total_break
                row[6].value = str(int(total_active))
                row[7].value = str(int(total_break))
                wb.save('data/work_log.xlsx')
            return
